from tempfile import NamedTemporaryFile

import re

from django.core.files.base import ContentFile
from django.http import HttpResponse
from django.utils.timezone import now

from openpyxl import Workbook
from openpyxl.worksheet.table import Table
from openpyxl.styles import Font, Border, Side

from swutils.string import transliterate

import background_tasks.models
import background_tasks.consts


class Excel:
    alphabet = [chr(code) for code in range(65, 91)]

    workbook_name = 'Отчет'
    headers = []

    head_static_sizes = {
        '№': 10,
    }

    def __init__(self, title=None, objects: list = None, background_task: background_tasks.models.Task = None):
        self.title = title
        self.objects = objects
        self.background_task = background_task

        self._workbook = None
        self._sheet = None
        self._row = 1

    def create_workbook(self):
        workbook = self.get_workbook()

        self.write_header()
        self.write_objects()

        file_name = transliterate(self.get_workbook_name(), space='_')
        # очистим имя файла от небуквенных символов
        file_name = re.sub("[^\\w]", "", file_name)
        file_name += '.xlsx'

        with NamedTemporaryFile() as tmp:
            workbook.save(tmp.name)
            tmp.seek(0)

            response = HttpResponse(content_type='application/ms-excel', content=tmp.read())
            response['Content-Disposition'] = 'attachment; filename=%s' % file_name

            if self.background_task:
                self.finish_background_task(workbook, response.content)

            return response

    @classmethod
    def create_workbook_background(cls, background_task: background_tasks.models.Task, **params):
        background_tasks.models.Task.objects.filter(id=background_task.id).update(
            status=background_tasks.consts.STATUS_IN_PROCESS,
            start_dt=now(),
            percent=0,
        )
        cls(background_task=background_task, **params['workbook_maker_kwargs']).create_workbook()

    def finish_background_task(self, workbook, content):
        background_tasks.models.Task.objects.filter(id=self.background_task.id).update(
            status=background_tasks.consts.STATUS_COMPLETE,
            finish_dt=now(),
            percent=100,
        )
        background_tasks.models.Log.objects.create(
            task=self.background_task,
            description='Формирование excel-файла завершено',
            status=background_tasks.consts.STATUS_COMPLETE,
        )

        file_name = f"{transliterate(self.get_workbook_name(), space='_')}.xls"
        background_task = background_tasks.models.Task.objects.get(id=self.background_task.id)
        workbook.save(file_name)
        background_task.result_attachment.save(file_name, ContentFile(content))

    def get_workbook(self):
        if not self._workbook:
            self._workbook = Workbook()
        return self._workbook

    def get_sheet(self):
        sheet = self.get_workbook().active
        sheet.title = self.get_sheet_name()
        return sheet

    def get_workbook_name(self):
        return self.workbook_name

    def get_sheet_name(self):
        return self.get_workbook_name()[:30]

    def get_headers(self):
        return list(self.get_head_static_sizes().keys()) or []

    def get_head_static_sizes(self):
        return self.head_static_sizes or {}

    def get_col_address(self, number: int) -> str:
        second_rank = number % len(self.alphabet)
        first_rank = number // len(self.alphabet)

        address = ''
        if first_rank:
            address = self.get_col_address(first_rank - 1)

        return address + self.alphabet[second_rank]

    def get_cell_address(self, row_number: int, col_number: int) -> str:
        """ Возвращает физический адрес ячейки типа А1 """
        return '%s%s' % (self.get_col_address(col_number), row_number + 1)

    def get_current_row(self):
        return self._row

    def add_current_row(self):
        self._row += 1

    def write_header(self):
        """ Заголовок над таблицей отчета """
        sheet = self.get_sheet()

        if not self.title:
            return

        title_cell = sheet.cell(self.get_current_row(), 1, self.title)

        bold_font = self.get_bold_font()
        self.set_font_size(bold_font, 14)
        title_cell.font = bold_font

        sheet.row_dimensions[self.get_current_row()].height = 30

        self.add_current_row()

    def write_objects(self):
        sheet = self.get_sheet()

        object_rows = self.get_object_rows()
        sheet.append(self.get_headers())

        for row in object_rows:
            sheet.append(row)

        current_row = self.get_current_row()

        for col_num, header in zip(range(1, len(self.get_headers())), self.get_headers()):
            cell = sheet[self.get_cell_address(current_row - 1, col_num)]
            cell.font = self.get_bold_font()

            sheet.column_dimensions[self.get_col_address(col_num - 1)].width = self.get_head_static_sizes().get(header, 30)

        cell_table_start = self.get_cell_address(current_row - 1, 1)
        cell_table_finish = self.get_cell_address(current_row - 1, len(self.get_headers()) - 1)

        tab = Table(
            displayName="Table1",
            ref=f"{cell_table_start}:{cell_table_finish}"
        )
        sheet.add_table(tab)

        for row in range(current_row, len(self.objects) + current_row + 1):
            for col in range(1, len(self.get_headers()) + 1):
                sheet.cell(row=row, column=col).border = self.get_border()

    def get_object_rows(self):
        raise NotImplementedError()

    def save_background_task_progress(self, percent, description):
        if not background_tasks.models.Task.objects.filter(id=self.background_task.id, percent=percent).exists():
            background_tasks.models.Log.objects.create(
                task=self.background_task,
                description=description
            )
            background_tasks.models.Task.objects.filter(id=self.background_task.id).update(percent=percent)

    @staticmethod
    def get_bold_font():
        return Font(bold=True)

    @staticmethod
    def get_border():
        return Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    @staticmethod
    def set_font_size(font, size):
        font.sz = size
        return font
